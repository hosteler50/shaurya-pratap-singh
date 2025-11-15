from flask import Flask, render_template, request, redirect, url_for, jsonify, Response, session, send_from_directory, abort
from flask_session import Session
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import uuid
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import shutil

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'hostel-review-secret-key-change-in-prod')
app.config['SESSION_TYPE'] = 'filesystem'
Session(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
DATA_FILE = os.path.join(DATA_DIR, 'hostels.xlsx')
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'hosteler50@gmail.com')


def ensure_data_file():
    os.makedirs(DATA_DIR, exist_ok=True)
    uploads_dir = os.path.join(BASE_DIR, 'static', 'uploads')
    os.makedirs(uploads_dir, exist_ok=True)
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        hs = wb.active
        hs.title = 'Hostels'
        hs.append(['id', 'name', 'location', 'description', 'image'])
        hs.append([str(uuid.uuid4()), 'Maple Residency', 'Downtown', 'Clean rooms, friendly staff', ''])
        hs.append([str(uuid.uuid4()), 'Seaside Lodge', 'Beachfront', 'Great view and lively area', ''])

        rs = wb.create_sheet('Reviews')
        rs.append(['hostel_id', 'reviewer_id', 'reviewer_name', 'reviewer_mobile', 'reviewer_college', 'reviewer_course', 'reviewer_address', 'rating_overall', 'rating_food', 'rating_cleaning', 'rating_staff', 'rating_location', 'rating_owner', 'fees_per_year', 'room_sharing', 'comment', 'date'])

        us = wb.create_sheet('Users')
        us.append(['id', 'email', 'password_hash', 'name'])

        wb.save(DATA_FILE)


def load_workbook_safe():
    ensure_data_file()
    wb = load_workbook(DATA_FILE)
    modified = False
    if 'Hostels' not in wb.sheetnames:
        hs = wb.create_sheet('Hostels')
        hs.append(['id', 'name', 'location', 'description', 'image'])
        modified = True
    if 'Reviews' not in wb.sheetnames:
        rs = wb.create_sheet('Reviews')
        rs.append(['hostel_id', 'reviewer_id', 'reviewer_name', 'reviewer_mobile', 'reviewer_college', 'reviewer_course', 'reviewer_address', 'rating_overall', 'rating_food', 'rating_cleaning', 'rating_staff', 'rating_location', 'rating_owner', 'fees_per_year', 'room_sharing', 'comment', 'date'])
        modified = True
    if 'Users' not in wb.sheetnames:
        us = wb.create_sheet('Users')
        us.append(['id', 'email', 'password_hash', 'name'])
        modified = True
    if modified:
        wb.save(DATA_FILE)
    return wb


def load_hostels():
    wb = load_workbook_safe()
    if 'Hostels' not in wb.sheetnames:
        return []
    hs = wb['Hostels']
    hostels = []
    for row in hs.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        hostels.append({
            'id': row[0],
            'name': row[1],
            'location': row[2],
            'description': row[3] or '',
            'image': row[4] or ''
        })
    return hostels


def load_reviews():
    wb = load_workbook_safe()
    if 'Reviews' not in wb.sheetnames:
        return []
    rs = wb['Reviews']
    reviews = []
    for row in rs.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        try:
            def to_num(v):
                try:
                    return float(v) if v is not None and str(v).strip() != '' else None
                except Exception:
                    return None

            # New format (17 cols): hostel_id, reviewer_id, reviewer_name, mobile, college, course, address, overall, food, cleaning, staff, location, owner, fees, room_sharing, comment, date
            if len(row) >= 17:
                hostel_id = row[0]
                reviewer_id = row[1]
                reviewer_name = row[2] or 'Anonymous'
                reviewer_mobile = row[3] or ''
                reviewer_college = row[4] or ''
                reviewer_course = row[5] or ''
                reviewer_address = row[6] or ''
                rating_overall = to_num(row[7])
                rating_food = to_num(row[8])
                rating_cleaning = to_num(row[9])
                rating_staff = to_num(row[10])
                rating_location = to_num(row[11])
                rating_owner = to_num(row[12])
                fees_per_year = row[13] or ''
                room_sharing = row[14] or ''
                comment = row[15] or ''
                date_val = row[16]
            # Old format (11 cols): hostel_id, reviewer_id, reviewer_name, overall, food, cleaning, staff, location, owner, comment, date
            elif len(row) >= 11:
                hostel_id = row[0]
                reviewer_id = row[1]
                reviewer_name = row[2] or 'Anonymous'
                reviewer_mobile = ''
                reviewer_college = ''
                reviewer_course = ''
                reviewer_address = ''
                rating_overall = to_num(row[3])
                rating_food = to_num(row[4])
                rating_cleaning = to_num(row[5])
                rating_staff = to_num(row[6])
                rating_location = to_num(row[7])
                rating_owner = to_num(row[8])
                fees_per_year = ''
                room_sharing = ''
                comment = row[9] or ''
                date_val = row[10]
            else:
                continue

            reviews.append({
                'hostel_id': hostel_id,
                'reviewer_id': reviewer_id,
                'reviewer_name': reviewer_name,
                'reviewer_mobile': reviewer_mobile,
                'reviewer_college': reviewer_college,
                'reviewer_course': reviewer_course,
                'reviewer_address': reviewer_address,
                'rating_overall': rating_overall,
                'rating_food': rating_food,
                'rating_cleaning': rating_cleaning,
                'rating_staff': rating_staff,
                'rating_location': rating_location,
                'rating_owner': rating_owner,
                'fees_per_year': fees_per_year,
                'room_sharing': room_sharing,
                'comment': comment,
                'date': date_val
            })
        except Exception:
            continue
    return reviews


def add_hostel(name, location, description=''):
    wb = load_workbook_safe()
    hs = wb['Hostels']
    new_id = str(uuid.uuid4())
    hs.append([new_id, name, location, description, ''])
    wb.save(DATA_FILE)
    return new_id


def add_review(hostel_id, reviewer_id, reviewer_name, reviewer_mobile, reviewer_college, reviewer_course, reviewer_address, rating_overall, rating_food, rating_cleaning, rating_staff, rating_location, rating_owner, fees_per_year, room_sharing, comment):
    wb = load_workbook_safe()
    rs = wb['Reviews']
    now = datetime.utcnow().isoformat()
    rs.append([hostel_id, reviewer_id, reviewer_name, reviewer_mobile, reviewer_college, reviewer_course, reviewer_address, rating_overall, rating_food, rating_cleaning, rating_staff, rating_location, rating_owner, fees_per_year, room_sharing, comment, now])
    wb.save(DATA_FILE)


def load_users():
    wb = load_workbook_safe()
    if 'Users' not in wb.sheetnames:
        return []
    us = wb['Users']
    users = []
    for row in us.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        users.append({
            'id': row[0],
            'email': row[1],
            'password_hash': row[2],
            'name': row[3]
        })
    return users


def is_admin():
    """Simple admin check: session email matches ADMIN_EMAIL."""
    return session.get('user_email') and session.get('user_email').lower() == ADMIN_EMAIL.lower()


def migrate_reviews_in_wb():
    """Normalize Reviews sheet rows into 17-column format.
    Returns number of rows migrated.
    """
    wb = load_workbook_safe()
    if 'Reviews' not in wb.sheetnames:
        return 0
    rs = wb['Reviews']
    rows = list(rs.iter_rows(values_only=True))
    if not rows:
        return 0

    new_header = ['hostel_id', 'reviewer_id', 'reviewer_name', 'reviewer_mobile', 'reviewer_college', 'reviewer_course', 'reviewer_address', 'rating_overall', 'rating_food', 'rating_cleaning', 'rating_staff', 'rating_location', 'rating_owner', 'fees_per_year', 'room_sharing', 'comment', 'date']
    new_rows = []

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        row = list(row)
        # New rows already in full format
        if len(row) >= 17:
            nr = row[:17]
        # Legacy 11-column format: map known indices into new layout
        elif len(row) >= 11:
            # legacy: 0 hostel_id,1 reviewer_id,2 reviewer_name,3 overall,4 food,5 cleaning,6 staff,7 location,8 owner,9 comment,10 date
            nr = [
                row[0],  # hostel_id
                row[1],  # reviewer_id
                row[2] or 'Anonymous',
                '',     # reviewer_mobile
                '',     # reviewer_college
                '',     # reviewer_course
                '',     # reviewer_address
                row[3], # rating_overall
                row[4], # rating_food
                row[5], # rating_cleaning
                row[6], # rating_staff
                row[7], # rating_location
                row[8], # rating_owner
                '',     # fees_per_year
                '',     # room_sharing
                row[9], # comment
                row[10] # date
            ]
        else:
            # Skip malformed rows
            continue
        new_rows.append(nr)

    # backup existing sheet by renaming
    try:
        timestamp = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
        backup_name = f'Reviews_backup_{timestamp}'
        rs.title = backup_name
    except Exception:
        # if renaming fails, ignore and continue
        pass

    # create new Reviews sheet
    rs_new = wb.create_sheet('Reviews')
    rs_new.append(new_header)
    for nr in new_rows:
        rs_new.append(nr)
    wb.save(DATA_FILE)
    return len(new_rows)


def backup_workbook_file():
    """Create a file-copy backup of the workbook and return backup path."""
    os.makedirs(DATA_DIR, exist_ok=True)
    timestamp = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
    backup_name = f'hostels_backup_{timestamp}.xlsx'
    backup_path = os.path.join(DATA_DIR, backup_name)
    # copy if exists
    if os.path.exists(DATA_FILE):
        shutil.copy2(DATA_FILE, backup_path)
        return backup_path
    return None


def user_by_email(email):
    users = load_users()
    for u in users:
        if u['email'].lower() == email.lower():
            return u
    return None


def user_by_id(user_id):
    users = load_users()
    for u in users:
        if u['id'] == user_id:
            return u
    return None


def create_user(email, password, name):
    if user_by_email(email):
        return None
    wb = load_workbook_safe()
    us = wb['Users']
    new_id = str(uuid.uuid4())
    pwd_hash = generate_password_hash(password)
    us.append([new_id, email, pwd_hash, name])
    wb.save(DATA_FILE)
    return new_id


def save_hostel_image(file_storage):
    if not file_storage:
        return ''
    filename = secure_filename(file_storage.filename)
    if filename == '':
        return ''
    uploads_dir = os.path.join(BASE_DIR, 'static', 'uploads')
    os.makedirs(uploads_dir, exist_ok=True)
    unique_name = f"{uuid.uuid4().hex}_{filename}"
    path = os.path.join(uploads_dir, unique_name)
    file_storage.save(path)
    return f"/static/uploads/{unique_name}"


def average_rating_for(hostel_id):
    reviews = [r for r in load_reviews() if r['hostel_id'] == hostel_id and r.get('rating_overall') is not None]
    if not reviews:
        return None
    return round(sum(r.get('rating_overall', 0) for r in reviews) / len(reviews), 2)


def average_ratings_for(hostel_id):
    reviews = [r for r in load_reviews() if r['hostel_id'] == hostel_id]
    if not reviews:
        return {
            'overall': None, 'food': None, 'cleaning': None, 'staff': None, 'location': None, 'owner': None
        }
    def avg(key):
        vals = [r.get(key) for r in reviews if r.get(key) is not None]
        return round(sum(vals)/len(vals), 2) if vals else None

    return {
        'overall': avg('rating_overall'),
        'food': avg('rating_food'),
        'cleaning': avg('rating_cleaning'),
        'staff': avg('rating_staff'),
        'location': avg('rating_location'),
        'owner': avg('rating_owner')
    }


@app.route('/admin/reviews')
def admin_reviews():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    if not is_admin():
        return "Forbidden", 403

    # load hostels map for name lookup
    hostels_list = load_hostels()
    hostels_map = {h['id']: h['name'] for h in hostels_list}
    reviews = load_reviews()
    # render template with full mobiles (admin only)
    return render_template('admin_reviews.html', reviews=reviews, hostels_map=hostels_map)


@app.route('/admin/migrate_reviews')
def admin_migrate_reviews():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    if not is_admin():
        return "Forbidden", 403

    count = migrate_reviews_in_wb()
    return render_template('admin_migrate_result.html', count=count)


@app.route('/admin/backup_workbook')
def admin_backup_workbook():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    if not is_admin():
        return "Forbidden", 403

    backup_path = backup_workbook_file()
    if backup_path:
        fname = os.path.basename(backup_path)
        return render_template('admin_backup_result.html', backup_file=fname)
    else:
        return render_template('admin_backup_result.html', backup_file=None)


@app.route('/admin/backups')
def admin_backups():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    if not is_admin():
        return "Forbidden", 403

    os.makedirs(DATA_DIR, exist_ok=True)
    files = []
    for name in sorted(os.listdir(DATA_DIR), reverse=True):
        if not name.lower().startswith('hostels_backup_'):
            continue
        path = os.path.join(DATA_DIR, name)
        try:
            size = os.path.getsize(path)
            mtime = datetime.utcfromtimestamp(os.path.getmtime(path)).isoformat() + 'Z'
        except Exception:
            size = None
            mtime = None
        files.append({'name': name, 'size': size, 'mtime': mtime})

    return render_template('admin_backups.html', backups=files)


@app.route('/admin/backups/download/<path:filename>')
def admin_backups_download(filename):
    if not session.get('user_id'):
        return redirect(url_for('login'))
    if not is_admin():
        return "Forbidden", 403

    # prevent path traversal by allowing only files from DATA_DIR and matching backup prefix
    if not filename.startswith('hostels_backup_'):
        abort(404)
    safe_path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(safe_path):
        abort(404)
    return send_from_directory(DATA_DIR, filename, as_attachment=True)


@app.route('/admin/backups/restore/<path:filename>', methods=['GET', 'POST'])
def admin_backups_restore(filename):
    if not session.get('user_id'):
        return redirect(url_for('login'))
    if not is_admin():
        return "Forbidden", 403

    # validate filename
    if not filename.startswith('hostels_backup_'):
        abort(404)

    backup_full = os.path.join(DATA_DIR, filename)
    if not os.path.exists(backup_full):
        abort(404)

    if request.method == 'GET':
        # show confirmation page
        try:
            size = os.path.getsize(backup_full)
            mtime = datetime.utcfromtimestamp(os.path.getmtime(backup_full)).isoformat() + 'Z'
        except Exception:
            size = None
            mtime = None
        return render_template('admin_restore_confirm.html', backup_file=filename, size=size, mtime=mtime)

    # POST -> perform restore (safe: create a pre-restore backup first)
    pre_backup = backup_workbook_file()
    try:
        # copy selected backup to DATA_FILE
        shutil.copy2(backup_full, DATA_FILE)
        restored = True
    except Exception as e:
        restored = False
        error = str(e)

    return render_template('admin_restore_result.html', backup_file=filename, pre_backup=os.path.basename(pre_backup) if pre_backup else None, restored=restored, error=(error if not restored else None))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        user = user_by_email(email)
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['user_email'] = user['email']
            return redirect(url_for('hostels'))
        else:
            return render_template('login.html', error='Invalid email or password')
    return render_template('login.html')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        name = request.form.get('name', '').strip()
        
        if not email or not password or not name:
            return render_template('signup.html', error='All fields required')
        if password != confirm_password:
            return render_template('signup.html', error='Passwords do not match')
        if len(password) < 4:
            return render_template('signup.html', error='Password must be at least 4 characters')
        if user_by_email(email):
            return render_template('signup.html', error='Email already registered')
        
        user_id = create_user(email, password, name)
        session['user_id'] = user_id
        session['user_name'] = name
        session['user_email'] = email
        return redirect(url_for('hostels'))
    return render_template('signup.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/hostels')
def hostels():
    q = request.args.get('q', '').strip().lower()
    hostels_list = load_hostels()
    if q:
        hostels_list = [h for h in hostels_list if q in (h['name'] or '').lower() or q in (h['location'] or '').lower()]

    for h in hostels_list:
        h['avg_rating'] = average_rating_for(h['id'])
        h['reviews'] = [r for r in load_reviews() if r['hostel_id'] == h['id']]
        av = average_ratings_for(h['id'])
        h['rating_counts'] = {
            'overall': av['overall'],
            'food': av['food'],
            'cleaning': av['cleaning'],
            'staff': av['staff'],
            'location': av['location'],
            'owner': av['owner']
        }

    return render_template('hostels.html', hostels=hostels_list, query=request.args.get('q', ''), current_user=session.get('user_id'))


@app.route('/review')
def review_form():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    hostels_list = load_hostels()
    selected = request.args.get('hostel_id')
    return render_template('review.html', hostels=hostels_list, selected=selected)


@app.route('/submit_review', methods=['POST'])
def submit_review():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    hostel_id = request.form.get('hostel_id')
    new_hostel_name = request.form.get('new_hostel_name', '').strip()
    new_hostel_location = request.form.get('new_hostel_location', '').strip()
    
    # New review fields
    reviewer_mobile = request.form.get('reviewer_mobile', '').strip()
    reviewer_college = request.form.get('reviewer_college', '').strip()
    reviewer_course = request.form.get('reviewer_course', '').strip()
    reviewer_address = request.form.get('reviewer_address', '').strip()
    fees_per_year = request.form.get('fees_per_year', '').strip()
    room_sharing = request.form.get('room_sharing', '').strip()
    
    # Rating fields
    rating_overall = request.form.get('rating_overall')
    rating_food = request.form.get('rating_food')
    rating_cleaning = request.form.get('rating_cleaning')
    rating_staff = request.form.get('rating_staff')
    rating_location = request.form.get('rating_location')
    rating_owner = request.form.get('rating_owner')
    comment = request.form.get('comment', '').strip()

    if new_hostel_name:
        file = request.files.get('new_hostel_image')
        image_path = save_hostel_image(file)
        wb = load_workbook_safe()
        hs = wb['Hostels']
        new_id = str(uuid.uuid4())
        hs.append([new_id, new_hostel_name, new_hostel_location, '', image_path])
        wb.save(DATA_FILE)
        hostel_id = new_id

    def to_float(v):
        try:
            return float(v) if v is not None and str(v).strip() != '' else None
        except Exception:
            return None

    overall_val = to_float(rating_overall)
    food_val = to_float(rating_food)
    cleaning_val = to_float(rating_cleaning)
    staff_val = to_float(rating_staff)
    location_val = to_float(rating_location)
    owner_val = to_float(rating_owner)

    if hostel_id:
        user_id = session.get('user_id')
        user_name = session.get('user_name', 'Anonymous')
        add_review(hostel_id, user_id, user_name, reviewer_mobile, reviewer_college, reviewer_course, reviewer_address, overall_val, food_val, cleaning_val, staff_val, location_val, owner_val, fees_per_year, room_sharing, comment)

    return redirect(url_for('hostels'))


@app.route('/export_reviews')
def export_reviews():
    hostel_id = request.args.get('hostel_id')
    reviews_list = load_reviews()
    if hostel_id:
        reviews_list = [r for r in reviews_list if r['hostel_id'] == hostel_id]
    csv_lines = ['hostel_id,reviewer_id,reviewer_name,rating_overall,rating_food,rating_cleaning,rating_staff,rating_location,rating_owner,comment,date']
    for r in reviews_list:
        row = [
            r.get('hostel_id') or '',
            r.get('reviewer_id') or '',
            (r.get('reviewer_name') or '').replace('"','""'),
            str(r.get('rating_overall') or ''),
            str(r.get('rating_food') or ''),
            str(r.get('rating_cleaning') or ''),
            str(r.get('rating_staff') or ''),
            str(r.get('rating_location') or ''),
            str(r.get('rating_owner') or ''),
            (r.get('comment') or '').replace('"','""'),
            r.get('date') or ''
        ]
        csv_lines.append('"' + '","'.join(row) + '"')
    csv_text = '\n'.join(csv_lines)
    return Response(csv_text, mimetype='text/csv', headers={"Content-Disposition": "attachment; filename=reviews.csv"})


@app.route('/api/hostels')
def api_hostels():
    hostels_list = load_hostels()
    return jsonify(hostels_list)


@app.route('/health')
def health():
    """Health check endpoint for keeping Render warm."""
    return jsonify({'status': 'ok'}), 200


if __name__ == '__main__':
    ensure_data_file()
    app.run(debug=True)

