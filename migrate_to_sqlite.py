#!/usr/bin/env python
"""
Migration script: Excel (openpyxl) -> SQLite (SQLAlchemy)
Run this ONCE before deploying to convert existing data.
"""
import os
import sys
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash
from app import app, db
from models import User, Hostel, Review
from datetime import datetime

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data', 'hostels.xlsx')

def migrate():
    """Load data from Excel and insert into SQLite."""
    
    # Create all tables
    with app.app_context():
        db.create_all()
        print("✓ Created database tables")
        
        if not os.path.exists(DATA_FILE):
            print("⚠ No Excel file found. Skipping migration.")
            return
        
        wb = load_workbook(DATA_FILE)
        
        # Migrate Users
        if 'Users' in wb.sheetnames:
            us = wb['Users']
            for row in us.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                user = User(
                    id=row[0],
                    email=row[1],
                    password_hash=row[2],
                    name=row[3]
                )
                try:
                    db.session.add(user)
                except Exception as e:
                    print(f"⚠ Skipping user {row[1]}: {e}")
            db.session.commit()
            print("✓ Migrated users")
        
        # Migrate Hostels
        if 'Hostels' in wb.sheetnames:
            hs = wb['Hostels']
            for row in hs.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                hostel = Hostel(
                    id=row[0],
                    name=row[1],
                    location=row[2],
                    description=row[3] or '',
                    image=row[4] or ''
                )
                try:
                    db.session.add(hostel)
                except Exception as e:
                    print(f"⚠ Skipping hostel {row[1]}: {e}")
            db.session.commit()
            print("✓ Migrated hostels")
        
        # Migrate Reviews
        if 'Reviews' in wb.sheetnames:
            rs = wb['Reviews']
            for row in rs.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                def to_num(v):
                    try:
                        return float(v) if v is not None and str(v).strip() != '' else None
                    except Exception:
                        return None
                
                try:
                    if len(row) >= 11:
                        hostel_id, reviewer_id, reviewer_name = row[0], row[1], row[2]
                        rating_overall = to_num(row[3])
                        rating_food = to_num(row[4])
                        rating_cleaning = to_num(row[5])
                        rating_staff = to_num(row[6])
                        rating_location = to_num(row[7])
                        rating_owner = to_num(row[8])
                        comment = row[9] or ''
                        date_val = row[10] or datetime.utcnow()
                    elif len(row) >= 6:
                        hostel_id, reviewer_id, reviewer_name = row[0], row[1], row[2]
                        rating_overall = to_num(row[3])
                        rating_food = rating_cleaning = rating_staff = rating_location = rating_owner = None
                        comment = row[4] or ''
                        date_val = row[5] or datetime.utcnow()
                    elif len(row) >= 5:
                        hostel_id = row[0]
                        reviewer_id = None
                        reviewer_name = row[1] or 'Anonymous'
                        rating_overall = to_num(row[2])
                        rating_food = rating_cleaning = rating_staff = rating_location = rating_owner = None
                        comment = row[3] or ''
                        date_val = row[4] or datetime.utcnow()
                    else:
                        continue
                    
                    review = Review(
                        hostel_id=hostel_id,
                        reviewer_id=reviewer_id,
                        reviewer_name=reviewer_name or 'Anonymous',
                        rating_overall=rating_overall,
                        rating_food=rating_food,
                        rating_cleaning=rating_cleaning,
                        rating_staff=rating_staff,
                        rating_location=rating_location,
                        rating_owner=rating_owner,
                        comment=comment,
                        created_at=date_val if isinstance(date_val, datetime) else datetime.fromisoformat(str(date_val))
                    )
                    db.session.add(review)
                except Exception as e:
                    print(f"⚠ Skipping review for hostel {row[0]}: {e}")
            
            db.session.commit()
            print("✓ Migrated reviews")
        
        print("\n✅ Migration complete!")

if __name__ == '__main__':
    migrate()
